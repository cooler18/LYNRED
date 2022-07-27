import os
import random
import sys
import numpy as np
from tqdm import tqdm
import openpyxl
from FUSION.classes.Mask import Mask
from FUSION.classes.Metrics import *
from FUSION.script.image_management import name_generator

SCRIPT_DIR = os.path.dirname(os.path.dirname(os.path.abspath(__file__)))
sys.path.append(SCRIPT_DIR)
sys.path.append(os.path.dirname(SCRIPT_DIR))
sys.path.append(os.path.dirname((os.path.dirname(SCRIPT_DIR))))
from FUSION.tools.method_fusion import *
from FUSION.tools.registration_tools import *


###################################

def create_workbook(nb_weights, name_sheet, path=None, new=False):
    if path is None or new:
        workbook = openpyxl.Workbook()
    else:
        workbook = openpyxl.load_workbook(path)
    sheet = workbook.create_sheet(name_sheet)
    for nb in range(nb_weights):
        sheet.cell(1 + 15 * nb, 1).value = 'Metrics'
        sheet.merge_cells(start_row=2 + 15 * nb, start_column=1, end_row=3 + 15 * nb, end_column=1)
        sheet.cell(2 + 15 * nb, 1).value = 'Fusion Method'

        sheet.merge_cells(start_row=1 + 15 * nb, start_column=2, end_row=1 + 15 * nb, end_column=7)
        sheet.merge_cells(start_row=14 + 15 * nb, start_column=2, end_row=14 + 15 * nb, end_column=7)
        sheet.cell(1 + 15 * nb, 2).value = 'SSIM'

        sheet.merge_cells(start_row=1 + 15 * nb, start_column=8, end_row=1 + 15 * nb, end_column=13)
        sheet.merge_cells(start_row=14 + 15 * nb, start_column=8, end_row=14 + 15 * nb, end_column=13)
        sheet.cell(1 + 15 * nb, 8).value = 'NEC'

        sheet.merge_cells(start_row=1 + 15 * nb, start_column=14, end_row=1 + 15 * nb, end_column=19)
        sheet.merge_cells(start_row=14 + 15 * nb, start_column=14, end_row=14 + 15 * nb, end_column=19)
        sheet.cell(1 + 15 * nb, 14).value = 'NMI'

        sheet.merge_cells(start_row=1 + 15 * nb, start_column=20, end_row=1 + 15 * nb, end_column=25)
        sheet.merge_cells(start_row=14 + 15 * nb, start_column=20, end_row=14 + 15 * nb, end_column=25)
        sheet.cell(1 + 15 * nb, 20).value = 'PSNR'

        sheet.merge_cells(start_row=1 + 15 * nb, start_column=26, end_row=1 + 15 * nb, end_column=31)
        sheet.merge_cells(start_row=14 + 15 * nb, start_column=26, end_row=14 + 15 * nb, end_column=31)
        sheet.cell(1 + 15 * nb, 26).value = 'RMSE'

        for i in range(5):
            sheet.merge_cells(start_row=2 + 15 * nb, start_column=2 + i * 6, end_row=2 + 15 * nb, end_column=3 + i * 6)
            sheet.cell(2 + 15 * nb, 2 + i * 6).value = 'RGB'
            sheet.merge_cells(start_row=2 + 15 * nb, start_column=4 + i * 6, end_row=2 + 15 * nb, end_column=5 + i * 6)
            sheet.cell(2 + 15 * nb, 4 + i * 6).value = 'IR'
            sheet.merge_cells(start_row=2 + 15 * nb, start_column=6 + i * 6, end_row=2 + 15 * nb, end_column=7 + i * 6)
            sheet.cell(2 + 15 * nb, 6 + i * 6).value = 'TOT'
            for j in range(3):
                sheet.cell(3 + 15 * nb, 2 + i * 6 + j * 2).value = 'mean'
                sheet.cell(3 + 15 * nb, 3 + i * 6 + j * 2).value = 'Std'

        sheet.cell(4 + 15 * nb, 1).value = 'Alpha + L'
        sheet.cell(5 + 15 * nb, 1).value = 'Alpha'
        sheet.cell(6 + 15 * nb, 1).value = 'SSIM + L'
        sheet.cell(7 + 15 * nb, 1).value = "SSIM"
        sheet.cell(8 + 15 * nb, 1).value = "SSIM_Scale + L"
        sheet.cell(9 + 15 * nb, 1).value = "SSIM_Scale"
        sheet.cell(10 + 15 * nb, 1).value = 'Saliency_Scale + L'
        sheet.cell(11 + 15 * nb, 1).value = 'Saliency_Scale'
        sheet.cell(12 + 15 * nb, 1).value = 'Saliency_Gaussian + L'
        sheet.cell(13 + 15 * nb, 1).value = 'Saliency_Gaussian'
        sheet.cell(14 + 15 * nb, 1).value = 'Ref Value RGB/IR'

    return workbook


###################################
## PART 1 : Dictionnary generation
active = False
###################################
if active:
    Time = ['Day', 'Night']
    num = np.uint8(np.linspace(0, 99, 100))
    metrics = [Metric_ssim, Metric_nec, Metric_nmi, Metric_psnr, Metric_rmse]
    weightsRGB = [0.1, 0.25, 0.5, 1, 2, 5, 10, 20]
    Laplacian_pyrs = [True, False]

    ####################################
    output = {}
    for t in Time:
        output[t] = {}
        for Laplacian_pyr in Laplacian_pyrs:
            output[t][Laplacian_pyr] = {}
            if Laplacian_pyr:
                l = 'laplacian'
            else:
                l = 'no_laplacian'
            for weightRGB in weightsRGB:
                output[t][Laplacian_pyr][weightRGB] = {}
                for n in tqdm(num):
                    output[t][Laplacian_pyr][weightRGB][n] = {}
                    number = name_generator(n)
                    imgL = ImageCustom(
                        "/home/godeta/PycharmProjects/LYNRED/LynredDataset/" + t + "/hybrid/infrared_projected/left" + number + ".png")
                    imgR = ImageCustom(
                        "/home/godeta/PycharmProjects/LYNRED/LynredDataset/" + t + "/hybrid/right/right" + number + ".png")

                    fus = imgR.LAB()
                    fus2 = imgR.LAB()
                    fus3 = imgR.LAB()
                    fus4 = imgR.LAB()
                    ref = imgR.LAB()

                    ##################################################
                    mask = Mask(imgR, imgL)
                    mask_ssim = mask.ssim(weightRGB=weightRGB, win_size=3)
                    mask_ssim_saliency = mask.saliency_ssim(verbose=False, level_max=2, weightRGB=weightRGB)
                    mask_saliency_gaussian = mask.saliency_gaussian(verbose=False, weightRGB=weightRGB, intensityRBG=1,
                                                                    intensityIR=2,
                                                                    colorRGB=2,
                                                                    edgeRGB=1, edgeIR=1)
                    mask_saliency_scale = mask.saliency_scale(verbose=False, weightRGB=weightRGB, intensityRBG=1,
                                                              intensityIR=2, colorRGB=2,
                                                              edgeRGB=1, edgeIR=1)
                    if Laplacian_pyr:
                        fus[:, :, 0] = laplacian_pyr_fusion(fus[:, :, 0], imgL, mask_ssim, octave=4)
                        fus2[:, :, 0] = laplacian_pyr_fusion(fus2[:, :, 0], imgL, mask_saliency_gaussian, octave=4)
                        fus3[:, :, 0] = laplacian_pyr_fusion(fus3[:, :, 0], imgL, mask_saliency_scale, octave=4)
                        fus4[:, :, 0] = laplacian_pyr_fusion(fus4[:, :, 0], imgL, mask_ssim_saliency, octave=4)
                        ref[:, :, 0] = laplacian_pyr_fusion(ref[:, :, 0], imgL,
                                                            np.ones_like(imgL) * (weightRGB / (1 + weightRGB)),
                                                            octave=4)
                    else:
                        fus[:, :, 0] = fus[:, :, 0] * mask_ssim + imgL * (1 - mask_ssim)
                        fus2[:, :, 0] = fus2[:, :, 0] * mask_saliency_gaussian + imgL * (1 - mask_saliency_gaussian)
                        fus3[:, :, 0] = fus3[:, :, 0] * mask_saliency_scale + imgL * (1 - mask_saliency_scale)
                        fus4[:, :, 0] = fus4[:, :, 0] * mask_ssim_saliency + imgL * (1 - mask_ssim_saliency)
                        ref[:, :, 0] = ImageCustom(
                            ref[:, :, 0] * np.ones_like(imgL) * (weightRGB / (1 + weightRGB)) + imgL * np.ones_like(
                                imgL) * (1 / (1 + weightRGB)))
                    for metric in metrics:
                        ############################ REFERENCE VALUE ###################################
                        reference = metric(imgR, imgL)
                        output[t][Laplacian_pyr][weightRGB][n][reference.metric] = {'ref': reference.value}
                        ############################ SSIM ###################################
                        output[t][Laplacian_pyr][weightRGB][n][reference.metric]["mask_ssim"] = {
                            "rgb": metric(imgR, fus[:, :, 0]).value,
                            "ir": metric(imgL, fus[:, :, 0]).value,
                            "tot": metric(imgR, imgL, fus).value}
                        ############################ SSIM SCALE ##############################
                        output[t][Laplacian_pyr][weightRGB][n][reference.metric]["mask_ssim_scale"] = {
                            "rgb": metric(imgR, fus4[:, :, 0]).value,
                            "ir": metric(imgL, fus4[:, :, 0]).value,
                            "tot": metric(imgR, imgL, fus4).value}
                        ####################### SALIENCY GAUSSIAN ##############################
                        output[t][Laplacian_pyr][weightRGB][n][reference.metric]["mask_sal_gaussian"] = {
                            "rgb": metric(imgR, fus2[:, :, 0]).value,
                            "ir": metric(imgL, fus2[:, :, 0]).value,
                            "tot": metric(imgR, imgL, fus2).value}
                        ######################### SALIENCY SCALE ##############################
                        output[t][Laplacian_pyr][weightRGB][n][reference.metric]["mask_sal_scale"] = {
                            "rgb": metric(imgR, fus3[:, :, 0]).value,
                            "ir": metric(imgL, fus3[:, :, 0]).value,
                            "tot": metric(imgR, imgL, fus3).value}
                        ############################# REF 1/2 ##############################
                        output[t][Laplacian_pyr][weightRGB][n][reference.metric]["1/2"] = {
                            "rgb": metric(imgR, ref[:, :, 0]).value,
                            "ir": metric(imgL, ref[:, :, 0]).value,
                            "tot": metric(imgR, imgL, ref).value}
                name = 'result_' + t + '_' + l + '_' + str(weightRGB)
                dict_out = output[t][Laplacian_pyr][weightRGB]
                with open('/home/godeta/PycharmProjects/LYNRED/FUSION/results/' + name, "wb") as f:
                    pickle.dump(dict_out, f)
                # print(dict_out)
###################################
## PART 2 : Read and plot Dictionnary
###################################
else:
    Time = ['Night_']
    num = np.uint8(np.linspace(0, 99, 100))
    imgL = ImageCustom(
        "/home/godeta/PycharmProjects/LYNRED/LynredDataset/" + Time[0][
                                                               :-1] + "/hybrid/infrared_projected/left00000.png")
    imgR = ImageCustom(
        "/home/godeta/PycharmProjects/LYNRED/LynredDataset/" + Time[0][:-1] + "/hybrid/right/right00000.png")
    metrics = [Metric_ssim(imgL, imgR).metric, Metric_nec(imgL, imgR).metric, Metric_nmi(imgL, imgR).metric,
               Metric_psnr(imgL, imgR).metric, Metric_rmse(imgL, imgR).metric]
    weightRGB = [0.1, 0.25, 0.5, 1, 2, 5, 10, 20]
    Laplacian_pyr = ['laplacian_', 'no_laplacian_']
    name_sheet = 'result_night'
    path = '/home/godeta/Documents/results.xlsx'
    workbook = create_workbook(len(weightRGB), name_sheet, path=path, new=False)
    sheet = workbook[name_sheet]
    for i, Lap in enumerate(Laplacian_pyr):
        for j, m in enumerate(metrics):
            for w, weight in enumerate(weightRGB):
                ref = []
                ssim = {'rgb': [], 'ir': [], 'tot': []}
                ssim_scale = {'rgb': [], 'ir': [], 'tot': []}
                sale_gaussian = {'rgb': [], 'ir': [], 'tot': []}
                sale_scale = {'rgb': [], 'ir': [], 'tot': []}
                alpha = {'rgb': [], 'ir': [], 'tot': []}
                for t in Time:
                    name = 'result_' + t + Lap + str(weight)
                    with open('/home/godeta/PycharmProjects/LYNRED/FUSION/results/' + name, 'rb') as f:
                        output = pickle.load(f)
                    for n in num:
                        dic = output[n][m]
                        ref.append(dic["ref"])
                        for k in ssim.keys():
                            ssim[k].append(dic["mask_ssim"][k])
                            ssim_scale[k].append(dic["mask_ssim_scale"][k])
                            sale_gaussian[k].append(dic["mask_sal_gaussian"][k])
                            sale_scale[k].append(dic["mask_sal_scale"][k])
                            alpha[k].append(dic["1/2"][k])
                for k in ssim.keys():
                    ssim[k] = np.array(ssim[k])
                    ssim_scale[k] = np.array(ssim_scale[k])
                    sale_gaussian[k] = np.array(sale_gaussian[k])
                    sale_scale[k] = np.array(sale_scale[k])
                    alpha[k] = np.array(alpha[k])
                ref = np.array(ref)
                print(f'\n############# {m} // {Lap}#####################')
                print(f' REF : {np.mean(ref)}+-{np.std(ref)}')
                print(
                    f'alpha RGB : {np.mean(alpha["rgb"])}+-{np.std(alpha["rgb"])}, IR : {np.mean(alpha["ir"])}+-{np.std(alpha["ir"])}, TOT : {np.mean(alpha["tot"])}+-{np.std(alpha["tot"])}')
                print(
                    f'SSIM RGB : {np.mean(ssim["rgb"])}+-{np.std(ssim["rgb"])}, IR : {np.mean(ssim["ir"])}+-{np.std(ssim["ir"])}, TOT : {np.mean(ssim["tot"])}+-{np.std(ssim["tot"])}')
                print(
                    f'SSIM SCALE RGB : {np.mean(ssim_scale["rgb"])}+-{np.std(ssim_scale["rgb"])}, IR : {np.mean(ssim_scale["ir"])}+-{np.std(ssim_scale["ir"])}, TOT : {np.mean(ssim_scale["tot"])}+-{np.std(ssim_scale["tot"])}')
                print(
                    f'Saliency Gauss RGB : {np.mean(sale_gaussian["rgb"])}+-{np.std(sale_gaussian["rgb"])}, IR : {np.mean(sale_gaussian["ir"])}+-{np.std(sale_gaussian["ir"])}, TOT : {np.mean(sale_gaussian["tot"])}+-{np.std(sale_gaussian["tot"])}')
                print(
                    f'Saliency Scale RGB : {np.mean(sale_scale["rgb"])}+-{np.std(sale_scale["rgb"])}, IR : {np.mean(sale_scale["ir"])}+-{np.std(sale_scale["ir"])}, TOT : {np.mean(sale_scale["tot"])}+-{np.std(sale_scale["tot"])}')
                sheet.cell(14 + 15 * w, 2 + 6 * j).value = ref.mean()
                for mask, item in enumerate([alpha, ssim, ssim_scale, sale_scale, sale_gaussian]):
                    for k, color in enumerate(['rgb', 'ir', 'tot']):
                        sheet.cell(4 + i + 2 * mask + 15 * w, 2 + 6 * j + 2 * k).value = item[color].mean()
                        sheet.cell(4 + i + 2 * mask + 15 * w, 3 + 6 * j + 2 * k).value = item[color].std()
        workbook.save(path)
    workbook.close()
